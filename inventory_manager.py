from tkinter import *
import openpyxl
from openpyxl import load_workbook
import glob
from tkinter import ttk
from datetime import datetime
from openpyxl.styles import Alignment
from openpyxl.styles import Color, PatternFill, Font, Border


root = Tk()
# root.option_add('*Font', 'Helvetica')
# root.configure(bg="teal")
# root.iconbitmap(default='favicon.ico')
root.title("INVENTORY MANAGER")
root.geometry("900x500")

redFill = PatternFill(start_color='ff6666', end_color='ff6666', fill_type='solid')
greenFill = PatternFill(start_color='5cd65c', end_color='5cd65c', fill_type='solid')
sheets = []
files = []
for filename in glob.glob('*.xlsx'):
    files.append(filename[:-5])
    files.sort()

if 'Export' not in files:
    create_wb_export = openpyxl.Workbook()
    create_sheet_export = create_wb_export['Sheet']
    create_export_column_names = ['A', 'B', 'C', 'D']
    create_export_column_size = [12, 30, 30, 12]
    for i in range(4):
        create_sheet_export.column_dimensions[create_export_column_names[i]].width = create_export_column_size[i]
    create_wb_export.save(filename='Export.xlsx')


def sheet_create():
    if dd_create_sheet.get() != 'Export':
        filename_for_sheet = dd_create_sheet.get() + '.xlsx'
        wb_temp = load_workbook(filename_for_sheet)
        opBal_temp = int(entry_opBal.get())
        sheet_list = wb_temp.sheetnames
        alreadyPresent = entry_sheet.get() in sheet_list
        if entry_sheet.get() != '' and dd_create_sheet.get() != '':
            if not alreadyPresent:
                wb_temp.create_sheet(entry_sheet.get())
        sheet_temp = wb_temp[entry_sheet.get()]
        column_names = ['A', 'B', 'C', 'D', 'E', 'F']
        column_size = [12, 12, 12, 12, 20, 40]
        column_headers = ['Date', 'Op. Bal.', 'Import', 'Export', 'Closing Bal.', 'Remarks']
        for i in range(6):
            sheet_temp.column_dimensions[column_names[i]].width = column_size[i]
        for i in range(6):
            cell_temp = column_names[i] + '1'
            sheet_temp[cell_temp].value = column_headers[i]
            sheet_temp[cell_temp].alignment = Alignment(horizontal='center')
        sheet_temp['B2'] = opBal_temp
        sheet_temp['C2'] = 0
        sheet_temp['D2'] = 0
        sheet_temp['E2'] = opBal_temp
        sheet_temp['A2'] = datetime.now().strftime("%d.%m.%Y")
        wb_temp.save(filename_for_sheet)


def file_create():
    temp = entry_file.get() + '.xlsx'
    if entry_file.get() != '' and entry_file.get() != 'Export':
        wb = openpyxl.Workbook()
        wb.save(filename=temp)
    files.append(entry_file.get())
    files.sort()
    dd_create_sheet.config(value=files)
    import_select_file.config(value=files)
    export_file1.config(value=files)
    export_file2.config(value=files)
    export_file3.config(value=files)
    export_file4.config(value=files)
    export_file5.config(value=files)


def sheets_refresh():
    global sheets
    filename_for_sheet = import_select_file.get() + '.xlsx'
    wb_temp = load_workbook(filename_for_sheet)
    sheets = wb_temp.sheetnames
    sheets.sort()
    import_select_sheet.config(value=sheets)


def import_item():
    import_file_name = (import_select_file.get() + '.xlsx')
    import_wb_temp = load_workbook(import_file_name)
    import_sheet_temp = import_wb_temp[import_select_sheet.get()]
    column_names = ['A', 'B', 'C', 'D', 'E', 'F']
    last_row = import_sheet_temp.max_row
    if import_remarks.get() == '':
        remarks = 'NIL'
    else:
        remarks = import_remarks.get()
    last_close_value = import_sheet_temp['E' + str(last_row)].value
    import_column_values = [datetime.now().strftime("%d.%m.%Y"), last_close_value, int(import_quantity.get()), 0,
                            int(import_quantity.get()) + int(last_close_value), remarks]
    for i in range(6):
        import_cell_temp = column_names[i] + str(last_row + 1)
        import_sheet_temp[import_cell_temp].value = import_column_values[i]
        import_sheet_temp[import_cell_temp].fill = greenFill
    import_wb_temp.save(filename=import_file_name)


def export_item(expo_title, expo_file, expo_sheet, expo_quantity):
    export_file_name = (expo_file + '.xlsx')
    export_wb_temp = load_workbook(export_file_name)
    export_sheet_temp = export_wb_temp[expo_sheet]
    column_names = ['A', 'B', 'C', 'D', 'E', 'F']
    last_row = export_sheet_temp.max_row
    remarks = expo_title
    last_close_value = export_sheet_temp['E' + str(last_row)].value
    export_column_values = [datetime.now().strftime("%d.%m.%Y"), last_close_value, 0, int(expo_quantity),
                            int(last_close_value) - int(expo_quantity), remarks]
    for i in range(6):
        export_cell_temp = column_names[i] + str(last_row + 1)
        export_sheet_temp[export_cell_temp].value = export_column_values[i]
        export_sheet_temp[export_cell_temp].fill = redFill
    export_wb_temp.save(filename=expo_file + '.xlsx')


def export_edit():
    export_total_list = []
    export_quantity_total = 0
    export_file_temp = load_workbook('Export.xlsx')
    export_sheet_temp = export_file_temp['Sheet']
    export_columns = ['A', 'B', 'C', 'D']
    export_last_row = export_sheet_temp.max_row
    start_row = export_last_row + 2
    export_sheet_temp['A' + str(start_row)].value = datetime.now().strftime("%d.%m.%Y")
    export_sheet_temp['B' + str(start_row)].value = export_title.get()

    export_sheet_temp['C' + str(start_row)].value = export_sheet1.get()
    export_sheet_temp['C' + str(start_row + 1)].value = export_sheet2.get()
    export_sheet_temp['C' + str(start_row + 2)].value = export_sheet3.get()
    export_sheet_temp['C' + str(start_row + 3)].value = export_sheet4.get()
    export_sheet_temp['C' + str(start_row + 4)].value = export_sheet5.get()

    if export_quantity1.get() != '' and export_quantity1.get() != 'Quantity':
        export_sheet_temp['D' + str(start_row)].value = int(export_quantity1.get())
        export_total_list.append(int(export_quantity1.get()))
        export_item(export_title.get(), export_file1.get(), export_sheet1.get(), export_quantity1.get())
    if export_quantity2.get() != '' and export_quantity2.get() != 'Quantity':
        export_sheet_temp['D' + str(start_row + 1)].value = int(export_quantity2.get())
        export_total_list.append(int(export_quantity2.get()))
        export_item(export_title.get(), export_file2.get(), export_sheet2.get(), export_quantity2.get())
    if export_quantity3.get() != '' and export_quantity3.get() != 'Quantity':
        export_sheet_temp['D' + str(start_row + 2)].value = int(export_quantity3.get())
        export_total_list.append(int(export_quantity3.get()))
        export_item(export_title.get(), export_file3.get(), export_sheet3.get(), export_quantity3.get())
    if export_quantity4.get() != '' and export_quantity4.get() != 'Quantity':
        export_sheet_temp['D' + str(start_row + 3)].value = int(export_quantity4.get())
        export_total_list.append(int(export_quantity4.get()))
        export_item(export_title.get(), export_file4.get(), export_sheet4.get(), export_quantity4.get())
    if export_quantity5.get() != '' and export_quantity5.get() != 'Quantity':
        export_sheet_temp['D' + str(start_row + 4)].value = int(export_quantity5.get())
        export_total_list.append(int(export_quantity5.get()))
        export_item(export_title.get(), export_file5.get(), export_sheet5.get(), export_quantity5.get())

    export_quantity_total = sum(export_total_list)

    export_sheet_temp['B' + str(start_row + 5)].value = "TOTAL"
    export_sheet_temp['D' + str(start_row + 5)].value = export_quantity_total
    export_file_temp.save(filename='Export.xlsx')


def sheets_refresh_export(file, sheet):
    global sheets
    export_filename_for_sheet = file.get() + '.xlsx'
    export_wb_temp = load_workbook(export_filename_for_sheet)
    export_sheets = export_wb_temp.sheetnames
    export_sheets.sort()
    sheet.config(value=export_sheets)


def file_create_popup():
    file_create_window = Tk()
    file_create_window.eval('tk::PlaceWindow . center')
    file_details_text = Label(file_create_window, text=f'Create a new file named {(entry_file.get()).upper()}?')
    file_details_text.pack()
    file_details_confirm = Button(file_create_window, text='Confirm', command=lambda: [file_create(), file_create_window.destroy(), entry_file.delete(0, END)])
    file_details_confirm.pack()
    file_create_window.mainloop()


def sheet_create_popup():
    sheet_create_window = Tk()
    sheet_create_window.eval('tk::PlaceWindow . center')
    sheet_details_text1 = Label(sheet_create_window, text=f'File name: {(dd_create_sheet.get()).upper()}')
    sheet_details_text1.grid(row=0, column=0)
    sheet_details_text2 = Label(sheet_create_window, text=f'Sheet name: {(entry_sheet.get()).upper()}')
    sheet_details_text2.grid(row=1, column=0)
    sheet_details_text3 = Label(sheet_create_window, text=f'Opening Balance: {entry_opBal.get()}')
    sheet_details_text3.grid(row=2, column=0)
    sheet_details_confirm = Button(sheet_create_window, text='Confirm', command=lambda: [sheet_create(), sheet_create_window.destroy(), entry_sheet.delete(0, END), entry_opBal.delete(0, END), dd_create_sheet.set('')])
    sheet_details_confirm.grid(row=3, column=0)
    sheet_create_window.mainloop()


def import_item_popup():
    import_item_window = Tk()
    if import_remarks.get() == '' or import_remarks.get() == 'Remarks':
        remarks = "NIL"
    else:
        remarks = import_remarks.get()
    import_item_window.eval('tk::PlaceWindow . center')
    import_popup_text1 = Label(import_item_window, text=f'File name: {(import_select_file.get()).upper()}')
    import_popup_text1.grid(row=0, column=0)
    import_popup_text2 = Label(import_item_window, text=f'Sheet name: {(import_select_sheet.get()).upper()}')
    import_popup_text2.grid(row=1, column=0)
    import_popup_text3 = Label(import_item_window, text=f'Import quantity: {import_quantity.get()}')
    import_popup_text3.grid(row=2, column=0)
    import_popup_text4 = Label(import_item_window, text=f'Remarks: {remarks}')
    import_popup_text4.grid(row=3, column=0)
    import_popup_confirm = Button(import_item_window, text='Confirm',
                                   command=lambda: [import_item(), import_item_window.destroy(), import_quantity.delete(0, END),
                                                    import_select_sheet.set(''), import_select_file.set(''), import_remarks.delete(0, END)])
    import_popup_confirm.grid(row=4, column=0)
    import_item_window.mainloop()


def export_edit_popup():
    export_item_window = Tk()
    export_item_window.eval('tk::PlaceWindow . center')
    export_popup_text1 = Label(export_item_window, text=f'Title: {export_title.get()}')
    export_popup_text1.grid(row=0, column=0)
    export_popup_text2 = Label(export_item_window, text=f'{export_file1.get()} {export_sheet1.get()} {export_quantity1.get()}')
    export_popup_text2.grid(row=1, column=0)
    export_popup_text3 = Label(export_item_window, text=f'{export_file2.get()} {export_sheet2.get()} {export_quantity2.get()}')
    export_popup_text3.grid(row=2, column=0)
    export_popup_text4 = Label(export_item_window, text=f'{export_file3.get()} {export_sheet3.get()} {export_quantity3.get()}')
    export_popup_text4.grid(row=3, column=0)
    export_popup_text5 = Label(export_item_window, text=f'{export_file4.get()} {export_sheet4.get()} {export_quantity4.get()}')
    export_popup_text5.grid(row=4, column=0)
    export_popup_text6 = Label(export_item_window, text=f'{export_file5.get()} {export_sheet5.get()} {export_quantity5.get()}')
    export_popup_text6.grid(row=5, column=0)
    export_popup_confirm = Button(export_item_window, text='Confirm',
                                  command=lambda: [export_edit(), export_item_window.destroy(), export_quantity1.delete(0, END),
                                                   export_quantity2.delete(0, END), export_quantity3.delete(0, END),
                                                   export_quantity4.delete(0, END), export_quantity5.delete(0, END),
                                                   export_title.delete(0, END)])
    export_popup_confirm.grid(row=6, column=0)
    export_item_window.mainloop()


def clear_entry(event, entry):
    entry.delete(0, END)
    entry.unbind('<Button-1>')


# CREATE FILE DESIGN AHEAD
label_file_create = Label(root, text='Create new File ')
label_file_create.grid(row=0, column=0, pady=20)

entry_file = Entry(root, width=30, border=1)
entry_file.grid(row=0, column=1)
entry_file.insert(0, 'New File name')
entry_file.bind("<Button-1>", lambda event: clear_entry(event, entry_file))

create_button1 = Button(root, command=file_create_popup, text='CREATE (file)', border=1)
create_button1.grid(row=0, column=2)

# CREATE SHEET DESIGN AHEAD
label_sheet_create = Label(root, text='Create new sheet  ')
label_sheet_create.grid(row=1, column=0, pady=20)

dd_create_sheet = ttk.Combobox(root, value=files)
dd_create_sheet['state'] = 'readonly'
dd_create_sheet.grid(row=1, column=1)

entry_sheet = Entry(root, width=30)
entry_sheet.grid(row=1, column=2)
entry_sheet.insert(0, 'New Sheet name')
entry_sheet.bind("<Button-1>", lambda event: clear_entry(event, entry_sheet))


entry_opBal = Entry(root, width=30)
entry_opBal.grid(row=1, column=3)
entry_opBal.insert(0, 'Opening quantity')
entry_opBal.bind("<Button-1>", lambda event: clear_entry(event, entry_opBal))

create_button2 = Button(root, text='CREATE (sheet)', command=sheet_create_popup)
create_button2.grid(row=1, column=4)

# IMPORT DESIGN AHEAD
label_import = Label(root, text='Import ')
label_import.grid(row=2, column=0, pady=20)

import_select_file = ttk.Combobox(root, value=files)
import_select_file['state'] = 'readonly'
import_select_file.grid(row=2, column=1)
# dd_select_file.current(0)

import_select_sheet = ttk.Combobox(root, value=sheets, postcommand=sheets_refresh)
import_select_sheet['state'] = 'readonly'
import_select_sheet.grid(row=2, column=2)

import_quantity = Entry(root, width=30)
import_quantity.grid(row=2, column=3)
import_quantity.insert(0, 'Quantity')
import_quantity.bind("<Button-1>", lambda event: clear_entry(event, import_quantity))

import_remarks = Entry(root, width=30)
import_remarks.grid(row=2, column=4)
import_remarks.insert(0, 'Remarks')
import_remarks.bind("<Button-1>", lambda event: clear_entry(event, import_remarks))

import_button3 = Button(root, text='IMPORT', command=import_item_popup)
import_button3.grid(row=2, column=5)

# EXPORT DESIGN AHEAD
export_label = Label(root, text='Export ')
export_label.grid(row=3, column=0, pady=20)

export_title = Entry(root, width=30)
export_title.grid(row=3, column=1)
export_title.insert(0, 'Export Title')
export_title.bind("<Button-1>", lambda event: clear_entry(event, export_title))

export_file1 = ttk.Combobox(root, value=files)
export_file1['state'] = 'readonly'
export_file1.grid(row=4, column=2)

export_sheet1 = ttk.Combobox(root, value=sheets, postcommand=lambda: sheets_refresh_export(export_file1, export_sheet1))
export_sheet1['state'] = 'readonly'
export_sheet1.grid(row=4, column=3)

export_quantity1 = Entry(root, width=30)
export_quantity1.grid(row=4, column=4)
export_quantity1.insert(0, 'Quantity')
export_quantity1.bind("<Button-1>", lambda event: clear_entry(event, export_quantity1))

export_file2 = ttk.Combobox(root, value=files)
export_file2['state'] = 'readonly'
export_file2.grid(row=5, column=2)

export_sheet2 = ttk.Combobox(root, value=sheets, postcommand=lambda: sheets_refresh_export(export_file2, export_sheet2))
export_sheet2['state'] = 'readonly'
export_sheet2.grid(row=5, column=3)

export_quantity2 = Entry(root, width=30)
export_quantity2.grid(row=5, column=4)
export_quantity2.insert(0, 'Quantity')
export_quantity2.bind("<Button-1>", lambda event: clear_entry(event, export_quantity2))

export_file3 = ttk.Combobox(root, value=files)
export_file3['state'] = 'readonly'
export_file3.grid(row=6, column=2)

export_sheet3 = ttk.Combobox(root, value=sheets, postcommand=lambda: sheets_refresh_export(export_file3, export_sheet3))
export_sheet3['state'] = 'readonly'
export_sheet3.grid(row=6, column=3)

export_quantity3 = Entry(root, width=30)
export_quantity3.grid(row=6, column=4)
export_quantity3.insert(0, 'Quantity')
export_quantity3.bind("<Button-1>", lambda event: clear_entry(event, export_quantity3))

export_file4 = ttk.Combobox(root, value=files)
export_file4['state'] = 'readonly'
export_file4.grid(row=7, column=2)

export_sheet4 = ttk.Combobox(root, value=sheets, postcommand=lambda: sheets_refresh_export(export_file4, export_sheet4))
export_sheet4['state'] = 'readonly'
export_sheet4.grid(row=7, column=3)

export_quantity4 = Entry(root, width=30)
export_quantity4.grid(row=7, column=4)
export_quantity4.insert(0, 'Quantity')
export_quantity4.bind("<Button-1>", lambda event: clear_entry(event, export_quantity4))

export_file5 = ttk.Combobox(root, value=files)
export_file5['state'] = 'readonly'
export_file5.grid(row=8, column=2)

export_sheet5 = ttk.Combobox(root, value=sheets, postcommand=lambda: sheets_refresh_export(export_file5, export_sheet5))
export_sheet5['state'] = 'readonly'
export_sheet5.grid(row=8, column=3)

export_quantity5 = Entry(root, width=30)
export_quantity5.grid(row=8, column=4)
export_quantity5.insert(0, 'Quantity')
export_quantity5.bind("<Button-1>", lambda event: clear_entry(event, export_quantity5))

export_button = Button(root, text='Export', width=25, command=export_edit_popup)
export_button.grid(row=9, column=2)

root.mainloop()
